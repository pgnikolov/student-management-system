from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def delete_row(sheet: Worksheet, first_name_del: str, last_name_del: str):
    """
        Delete all rows from a worksheet in the given workbook.
    Args:
        sheet: The worksheet to delete from.
        first_name_del: Student's first name to delete.
        last_name_del: Student's last name to delete.
    Return:
        sheet without the deleted row.
    """

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        first_name = row[0].value
        last_name = row[1].value

        if first_name == first_name_del and last_name == last_name_del:
            sheet.delete_rows(row[0].row)
            break

    return sheet


def add_student(student: list, wb_year: Workbook, class_students: str):
    """
    Add a new student record to woorkbook.

    Args:
        student (list): Student info
        wb_year (Workbook): The workbook to add the student.
        class_students (str): The class to which  student will be added

    Return:
        None
    """
    all_students_info = wb_year.worksheets[0]
    group_sheet = wb_year[class_students]

    all_students_info.append(student)
    group_sheet.append(student)
    return wb_year


def delete_student(first_name_del: str, last_name_del: str, class_students: str, wb_year: Workbook):
    """
        Delte student record from woorkbook(from all students info and from group)
    Args:
        first_name_del (str): Student's last name.
        last_name_del (str): Student's last name.
        class_students (str): The class to which student belongs.
        wb_year (Workbook): The workbook to delete the student from.
    Return:
        wb_year (Workbook): Whithout the student has been deleted.
    """
    all_students_info = wb_year.worksheets[0]
    group_sheet = wb_year[class_students]

    delete_row(all_students_info, first_name_del, last_name_del)
    delete_row(group_sheet, first_name_del, last_name_del)

    return wb_year


def search_student(first_name_search: str, last_name_search: str, wb_year: Workbook):
    """
        This function searches for a student in the students Excel file by name.
    Args:
        first_name_search (str): Student's first name.
        last_name_search (str): Student's last name.
        wb_year (Workbook): The workbook to search the student.
    return
        None
    """
    serch_sheet = wb_year.worksheets[0]

    rows = serch_sheet.iter_rows(min_row=2, min_col=1, max_col=2)

    for fn, ln in rows:
        if fn.value == first_name_search and ln.value == last_name_search:
            return f"Row number: {fn.row}, Name {fn.value} {ln.value},"

    return None


def list_all_students_by_group(file_paht: str, class_students: str):
    """
        Print all students who are in same group (class).
    Args:
        file_paht (str): The file path of the Excel file.
        class_students (str): The group(class) students we want.
    Return:
        None
    """
    wb = load_workbook(file_paht)
    sheet = wb[class_students]

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        print(*(cell.value for cell in row))


def list_all_students_by_year(file_paht: str):
    """
        Print all students from same year.
    Args:
        file_paht (str): The file path of the Excel file.
    Return:
        None
    """
    wb = load_workbook(file_paht)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        print(*(cell.value for cell in row))


def new_all_students_custom(file_name: str, sheet_name: str, columns: int):
    """
        Create Excel file and ask the user for amount of columns and their names.
    Args:
        file_name (str): Name of file (example "all_2024" all students who start 2024)
        sheet_name (str): Name of sheet in Excel workbook
        columns (int): Number of columns in Sheet
    Return:
         None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    for i in range(1, columns + 1):
        sheet.cell(row=1, column=i).value = input("Enter column title: ")

    wb.save(filename=file_name)


def new_all_students_default(file_name: str, sheet_name: str):
    """
    Creates a new Excel file with a default structure for adding student information.

    Args:
        file_name (str): The name of the file (for example - "all_2024.xlsx").
        sheet_name (str): The name of the sheet in the Excel workbook.

    Returns:
        None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    sheet['A1'] = 'first_name'
    sheet['B1'] = 'last_name'
    sheet['C1'] = 'gender'
    sheet['D1'] = 'date_of_birth'
    sheet['E1'] = "parent"

    wb.save(filename=file_name)


def main():
    """
    Main function to provide user interaction.
    """

    while True:
        # Display menu options
        print("\nStudent Management System")
        print("1. Add Student")
        print("2. Delete Student")
        print("3. Search Student")
        print("4. List All Students in same Group")
        print("5. List All Students in same Year")
        print("6. Create new custom file")
        print("7. Create new default file")
        print("8. Exit")

        # Prompt user for their choice
        choice = input("Enter your choice: ")

        if choice == '1':
            new_student = [input("First name: ").capitalize(), input("Last name: ").capitalize(),
                           input("Gender (Male/Female): ").capitalize(), input("Date of birth(YYYY-MM-DD): "),
                           input("Parent First and Last name together: ").capitalize(), ]
            students_group = input("Choose a group ('A', 'B', 'C' or 'D'): ")
            new_student.append(new_student)
            path_to_wb = input("Enter the path to a file, where you want to add new student: ")
            wb_to_add = load_workbook(path_to_wb)
            add_student(new_student, wb_to_add, students_group)
            wb_to_add.save(filename=path_to_wb)
        elif choice == '2':
            first_name = input("Enter the first name of the student to update: ").capitalize()
            last_name = input("Enter the last name of the student to update: ").capitalize()
            group_name = input("Enter the in which group is the student('A', 'B', 'C' or 'D'): ").capitalize()
            file_name = input("Enter path to the file: ")
            wb_remove_student = load_workbook(file_name)
            delete_student(first_name, last_name, group_name, wb_remove_student)
            wb_remove_student.save(filename=file_name)
        elif choice == '3':
            first_name = input("Enter the first name of the student to update: ").capitalize()
            last_name = input("Enter the last name of the student to update: ").capitalize()
            file_name = input("Enter path to the file: ")
            wb_to_search = load_workbook(file_name)
            search_student(first_name, last_name, wb_to_search)
        elif choice == '4':
            file_name = input("Enter path to the file: ").lower()
            group_name = input("Enter the name of the group('A', 'B', 'C', 'D': ").capitalize()
            list_all_students_by_group(file_name, group_name)
        elif choice == '5':
            file_name = input("Enter path to the file: ").lower()
            list_all_students_by_year(file_name)
        elif choice == "6":
            file_name = input("Enter path to the file: ").lower()
            sheet_name = input("Enter name for first sheet in your file: ")
            amount_of_columns = int(input("Enter the number of columns you need: "))
            new_all_students_custom(file_name, sheet_name, amount_of_columns)
        elif choice == '7':
            file_name = input("Enter path to the file: ").lower()
            sheet_name = input("Enter name for first sheet in your file: ")
            new_all_students_default(file_name, sheet_name)
        elif choice == "8":
            print("Exiting...")
            break
        else:
            print("Invalid choice, please try again.")


if __name__ == "__main__":
    main()
