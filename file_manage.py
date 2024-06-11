from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def delete_row(sheet: Worksheet, first_name_del: str, last_name_del: str):
    """
        Delete all rows from a worksheet in the given workbook.
    Args:
        sheet: The worksheet to delete from.
        first_name_del: Student's first name to delete.
        last_name_del: Student's last name to delete.
    Return:
        sheet without the deleted row.
    """

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        first_name = row[0].value
        last_name = row[1].value

        if first_name == first_name_del and last_name == last_name_del:
            # Delete the row if the first and last names match
            sheet.delete_rows(row[0].row)
            break  # Exit the loop after deleting the first matching row

    return sheet


def new_all_students_custom(sheet_name: str, file_name: str, columns: int):
    """
    Create Excel file and ask the user for amount of columns
    and their names.

    Args:
        sheet_name (str): Name of sheet in Excel workbook
        file_name (str): Name of file (example "all_2024" all students who start 2024)
        columns (int): Number of columns in Sheet

    Return:
         None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    for i in range(1, columns + 1):
        sheet.cell(row=1, column=i).value = input("Enter column title: ")

    wb.save(filename=file_name)


def new_all_students_default(file_name: str, sheet_name: str):
    """
    Creates a new Excel file with a default structure for adding student information.

    Args:
        file_name (str): The name of the file (for example - "all_2024.xlsx").
        sheet_name (str): The name of the sheet in the Excel workbook.

    Returns:
        None
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    sheet['A1'] = 'first_name'
    sheet['B1'] = 'last_name'
    sheet['C1'] = 'gender'
    sheet['D1'] = 'date_of_birth'
    sheet['E1'] = "parent"

    wb.save(filename=file_name)


def load_existing_file(file_path: str):
    """
    Loads an existing Excel file and returns the workbook and its sheets.

    Args:
        file_path (str): The path to the Excel file to be loaded.
    Returns:
        workbook
    """
    workbook = load_workbook(file_path)

    return workbook


def save_existing_file(workbook, file_name: str):
    """
    Saves the given workbook to the specified file.

    Args:
        workbook (Workbook): The workbook to be saved.
        file_name (str): The name of the file to save the workbook to.

    Returns:
        None
    """
    workbook.save(filename=file_name)
    print(f"Workbook saved as {file_name}")
